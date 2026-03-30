"""Стилизация Excel-файлов через openpyxl."""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

HEADER_BG = "1F4E79"
ROW_ALT_BG = "DEEAF1"
WHITE = "FFFFFF"


def apply_header_style(ws: Worksheet, row: int = 1) -> None:
    """Применяет стиль заголовка к строке row."""
    header_fill = PatternFill(fill_type="solid", fgColor=HEADER_BG)
    header_font = Font(bold=True, color=WHITE, size=11)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for cell in ws[row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    ws.row_dimensions[row].height = 20


def apply_row_styles(ws: Worksheet, start_row: int = 2) -> None:
    """Чередующиеся строки + выравнивание."""
    alt_fill = PatternFill(fill_type="solid", fgColor=ROW_ALT_BG)
    white_fill = PatternFill(fill_type="solid", fgColor=WHITE)
    data_align = Alignment(vertical="center", wrap_text=False)

    for i, row in enumerate(ws.iter_rows(min_row=start_row)):
        fill = alt_fill if i % 2 == 0 else white_fill
        for cell in row:
            cell.fill = fill
            cell.alignment = data_align


def auto_width(ws: Worksheet, min_width: int = 10, max_width: int = 50) -> None:
    """Авто-ширина колонок по содержимому."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val = str(cell.value or "")
                if len(val) > max_len:
                    max_len = len(val)
            except Exception:
                pass
        width = max(min_width, min(max_len + 2, max_width))
        ws.column_dimensions[col_letter].width = width


def freeze_header(ws: Worksheet) -> None:
    ws.freeze_panes = "A2"


def format_amount_column(ws: Worksheet, col_index: int, start_row: int = 2) -> None:
    """Числовой формат для денег."""
    col_letter = get_column_letter(col_index)
    for row in ws.iter_rows(min_row=start_row, min_col=col_index, max_col=col_index):
        for cell in row:
            cell.number_format = '#,##0.00 "€"'


def format_date_column(ws: Worksheet, col_index: int, start_row: int = 2) -> None:
    """Формат даты."""
    col_letter = get_column_letter(col_index)
    for row in ws.iter_rows(min_row=start_row, min_col=col_index, max_col=col_index):
        for cell in row:
            cell.number_format = "DD.MM.YYYY"


def apply_total_row_style(ws: Worksheet, row: int) -> None:
    """Жирный шрифт для итоговой строки."""
    bold_font = Font(bold=True)
    total_fill = PatternFill(fill_type="solid", fgColor="BDD7EE")
    for cell in ws[row]:
        cell.font = bold_font
        cell.fill = total_fill
