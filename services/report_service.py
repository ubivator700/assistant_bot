"""Генерация Excel-отчётов через openpyxl."""
from __future__ import annotations

import io
import json
from datetime import datetime, timedelta
from typing import Optional

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference, PieChart
from openpyxl.chart.series import DataPoint

from services import db_service
from utils.excel_exporter import (
    apply_header_style,
    apply_row_styles,
    auto_width,
    freeze_header,
    format_amount_column,
    format_date_column,
    apply_total_row_style,
)


# ─── Отчёт по расходам ────────────────────────────────────────────────────────

async def generate_expense_report(
    user_id: int,
    date_from: datetime,
    date_to: datetime,
) -> io.BytesIO:
    expenses = await db_service.get_expenses(user_id, since=date_from, until=date_to)

    wb = Workbook()

    # ── Лист 1: Все расходы ──────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Расходы"

    headers = ["Дата", "Категория", "Описание", "Сумма", "Валюта"]
    ws1.append(headers)
    apply_header_style(ws1)

    total = 0.0
    for e in expenses:
        ws1.append([
            e.created_at.strftime("%d.%m.%Y"),
            e.category or "другое",
            e.description or "",
            float(e.amount),
            e.currency,
        ])
        total += float(e.amount)

    # Итоговая строка
    total_row = ws1.max_row + 1
    ws1.cell(row=total_row, column=1).value = "ИТОГО"
    ws1.cell(row=total_row, column=4).value = total
    apply_total_row_style(ws1, total_row)

    apply_row_styles(ws1)
    format_date_column(ws1, 1)
    format_amount_column(ws1, 4)
    auto_width(ws1)
    freeze_header(ws1)

    # ── Лист 2: По категориям ────────────────────────────────────────────────
    ws2 = wb.create_sheet("По категориям")
    ws2.append(["Категория", "Сумма (EUR)", "% от итога"])
    apply_header_style(ws2)

    by_cat: dict = {}
    for e in expenses:
        cat = e.category or "другое"
        by_cat[cat] = by_cat.get(cat, 0) + float(e.amount)

    cat_row_start = 2
    for cat, amount in sorted(by_cat.items(), key=lambda x: -x[1]):
        pct = round(amount / total * 100, 1) if total > 0 else 0
        ws2.append([cat, amount, pct])

    apply_row_styles(ws2)
    format_amount_column(ws2, 2)
    auto_width(ws2)
    freeze_header(ws2)

    # Диаграмма по категориям
    if by_cat:
        cat_count = len(by_cat)
        chart = BarChart()
        chart.type = "col"
        chart.title = "Расходы по категориям"
        chart.y_axis.title = "EUR"
        chart.x_axis.title = "Категория"
        chart.width = 20
        chart.height = 12

        data = Reference(ws2, min_col=2, min_row=1, max_row=1 + cat_count)
        cats = Reference(ws2, min_col=1, min_row=2, max_row=1 + cat_count)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws2.add_chart(chart, f"E2")

    # ── Лист 3: Динамика по дням ─────────────────────────────────────────────
    ws3 = wb.create_sheet("Динамика")
    ws3.append(["Дата", "Сумма за день (EUR)"])
    apply_header_style(ws3)

    by_day: dict = {}
    for e in expenses:
        day = e.created_at.strftime("%d.%m.%Y")
        by_day[day] = by_day.get(day, 0) + float(e.amount)

    for day, amount in sorted(by_day.items()):
        ws3.append([day, amount])

    apply_row_styles(ws3)
    format_amount_column(ws3, 2)
    auto_width(ws3)
    freeze_header(ws3)

    if by_day:
        line_chart = LineChart()
        line_chart.title = "Динамика расходов"
        line_chart.y_axis.title = "EUR"
        line_chart.x_axis.title = "Дата"
        line_chart.width = 20
        line_chart.height = 12
        day_count = len(by_day)
        data = Reference(ws3, min_col=2, min_row=1, max_row=1 + day_count)
        cats = Reference(ws3, min_col=1, min_row=2, max_row=1 + day_count)
        line_chart.add_data(data, titles_from_data=True)
        line_chart.set_categories(cats)
        ws3.add_chart(line_chart, "D2")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─── Отчёт по задачам ─────────────────────────────────────────────────────────

async def generate_tasks_report(
    user_id: int,
    date_from: datetime,
    date_to: datetime,
) -> io.BytesIO:
    all_tasks = await db_service.get_tasks(user_id, status=None)
    tasks_in_range = [
        t for t in all_tasks
        if t.created_at and date_from <= t.created_at <= date_to
    ]

    wb = Workbook()

    # ── Лист 1: Все задачи ───────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Все задачи"
    ws1.append(["Название", "Приоритет", "Статус", "Дедлайн", "Создана"])
    apply_header_style(ws1)

    prio_label = {1: "Высокий", 2: "Средний", 3: "Низкий"}
    status_label = {"pending": "В работе", "done": "Выполнено", "cancelled": "Отменено"}

    for t in tasks_in_range:
        ws1.append([
            t.title,
            prio_label.get(t.priority, str(t.priority)),
            status_label.get(t.status, t.status),
            t.deadline.strftime("%d.%m.%Y %H:%M") if t.deadline else "",
            t.created_at.strftime("%d.%m.%Y") if t.created_at else "",
        ])

    apply_row_styles(ws1)
    auto_width(ws1)
    freeze_header(ws1)

    # ── Лист 2: Статистика ───────────────────────────────────────────────────
    ws2 = wb.create_sheet("Статистика")
    ws2.append(["Статус", "Количество"])
    apply_header_style(ws2)

    done = sum(1 for t in tasks_in_range if t.status == "done")
    pending = sum(1 for t in tasks_in_range if t.status == "pending")
    cancelled = sum(1 for t in tasks_in_range if t.status == "cancelled")
    overdue = sum(
        1 for t in tasks_in_range
        if t.status == "pending" and t.deadline and t.deadline < datetime.now()
    )

    for label, count in [
        ("Выполнено", done),
        ("В работе", pending),
        ("Просрочено", overdue),
        ("Отменено", cancelled),
    ]:
        ws2.append([label, count])

    apply_row_styles(ws2)
    auto_width(ws2)
    freeze_header(ws2)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─── Еженедельный дайджест ────────────────────────────────────────────────────

async def generate_weekly_digest(user_id: int) -> io.BytesIO:
    now = datetime.now()
    week_start = now - timedelta(days=7)

    expenses = await db_service.get_expenses(user_id, since=week_start)
    tasks = await db_service.get_tasks(user_id, status=None)
    tasks_week = [t for t in tasks if t.created_at and t.created_at >= week_start]
    meetings = await db_service.get_meetings(user_id, since=week_start, until=now)

    wb = Workbook()

    # Лист 1: Расходы недели
    ws1 = wb.active
    ws1.title = "Расходы недели"
    ws1.append(["Дата", "Категория", "Описание", "Сумма", "Валюта"])
    apply_header_style(ws1)
    for e in expenses:
        ws1.append([
            e.created_at.strftime("%d.%m.%Y"),
            e.category or "другое",
            e.description or "",
            float(e.amount),
            e.currency,
        ])
    if expenses:
        total_row = ws1.max_row + 1
        ws1.cell(row=total_row, column=1).value = "ИТОГО"
        ws1.cell(row=total_row, column=4).value = sum(float(e.amount) for e in expenses)
        apply_total_row_style(ws1, total_row)
    apply_row_styles(ws1)
    format_amount_column(ws1, 4)
    auto_width(ws1)
    freeze_header(ws1)

    # Лист 2: Задачи недели
    ws2 = wb.create_sheet("Задачи недели")
    ws2.append(["Название", "Приоритет", "Статус", "Дедлайн"])
    apply_header_style(ws2)
    prio_label = {1: "Высокий 🔴", 2: "Средний 🟡", 3: "Низкий ⚪"}
    for t in tasks_week:
        ws2.append([
            t.title,
            prio_label.get(t.priority, "-"),
            "✅ Выполнено" if t.status == "done" else "⏳ В работе",
            t.deadline.strftime("%d.%m.%Y") if t.deadline else "",
        ])
    apply_row_styles(ws2)
    auto_width(ws2)
    freeze_header(ws2)

    # Лист 3: Встречи
    ws3 = wb.create_sheet("Встречи")
    ws3.append(["Дата/время", "Название", "Участники", "Место"])
    apply_header_style(ws3)
    for m in meetings:
        participants = ""
        if m.participants:
            try:
                parts = json.loads(m.participants)
                participants = ", ".join(parts)
            except Exception:
                participants = m.participants
        ws3.append([
            m.start_dt.strftime("%d.%m.%Y %H:%M"),
            m.title,
            participants,
            m.location or "",
        ])
    apply_row_styles(ws3)
    auto_width(ws3)
    freeze_header(ws3)

    # Лист 4: Итог (саммари от Claude)
    ws4 = wb.create_sheet("Итог")
    ws4.append(["Параметр", "Значение"])
    apply_header_style(ws4)

    total_exp = sum(float(e.amount) for e in expenses)
    daily_avg = total_exp / 7 if expenses else 0
    tasks_done = sum(1 for t in tasks_week if t.status == "done")

    summary_data = [
        ("Период", f"{week_start.strftime('%d.%m.%Y')} — {now.strftime('%d.%m.%Y')}"),
        ("Итого расходов", f"{total_exp:.2f} EUR"),
        ("Среднее в день", f"{daily_avg:.2f} EUR"),
        ("Задач создано", len(tasks_week)),
        ("Задач выполнено", tasks_done),
        ("Встреч", len(meetings)),
    ]

    try:
        from services.ai_service import answer_data_query
        context = (
            f"Расходы: {total_exp:.2f} EUR за 7 дней, {len(expenses)} операций\n"
            f"Задачи: создано {len(tasks_week)}, выполнено {tasks_done}\n"
            f"Встречи: {len(meetings)} встреч"
        )
        ai_summary = await answer_data_query(
            user_id,
            "Дай краткий итог недели в 3-4 предложениях",
            context,
        )
        summary_data.append(("Итог недели (AI)", ai_summary))
    except Exception:
        pass

    for row_data in summary_data:
        ws4.append(list(row_data))

    apply_row_styles(ws4)
    auto_width(ws4)
    freeze_header(ws4)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
